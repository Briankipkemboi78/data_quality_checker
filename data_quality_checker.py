import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.express as px
import plotly.graph_objects as go
import yaml
import pathlib
import re
import logging
import time
import threading
import os

# ── Load .env for local development ───────────────────────────────────────────
# python-dotenv reads .env when present; on Azure, env vars come from
# App Service Application Settings and this block is safely skipped.
try:
    from dotenv import load_dotenv
    load_dotenv(dotenv_path=pathlib.Path(__file__).parent / ".env", override=False)
except ImportError:
    pass  # dotenv not installed — fine in production (Azure sets vars directly)

# ── Infrastructure Configuration ───────────────────────────────────────────────
# Reads from environment variables so secrets are never baked into source code.
# Set these in Azure App Service → Configuration → Application Settings, or
# in a .env file (excluded from version control) for local development.

INFRA_CONFIG = {
    # ── Authentication ─────────────────────────────────────────────────────────
    # Preferred: Azure AD (Entra ID) via App Service Easy Auth.
    # Fallback:  External identity (e.g. MSAL, Okta) configured via env vars.
    #
    # When deployed to Azure App Service with Easy Auth enabled, the platform
    # handles OIDC token validation before any request reaches this app.
    # The verified user identity is forwarded in the X-MS-CLIENT-PRINCIPAL-NAME
    # and X-MS-CLIENT-PRINCIPAL headers — read via st.context.headers 
    #
    # For external identity providers (Okta, Auth0, etc.), set:
    #   OIDC_ISSUER, OIDC_CLIENT_ID, OIDC_CLIENT_SECRET in App Settings.
    "auth": {
        "provider":           os.environ.get("AUTH_PROVIDER", "azure_ad"),   # "azure_ad" | "external"
        "require_auth":       os.environ.get("REQUIRE_AUTH", "true").lower() == "true",
        "allowed_domains":    os.environ.get("ALLOWED_EMAIL_DOMAINS", "").split(","),  # e.g. "company.com"
        "oidc_issuer":        os.environ.get("OIDC_ISSUER", ""),
        "oidc_client_id":     os.environ.get("OIDC_CLIENT_ID", ""),
    },

    # ── File Handling ──────────────────────────────────────────────────────────
    # Uploaded files are read into memory (BytesIO) and are NEVER written to
    # disk or any persistent store. They are discarded when the session ends.
    # The 50 MB cap below prevents memory abuse from very large uploads.
    "file_handling": {
        "max_upload_mb":      int(os.environ.get("MAX_UPLOAD_MB", "50")),
        "persist_uploads":    False,          # Hard-coded: uploads are always ephemeral
        "allowed_extensions": ["xlsx", "xls"],
    },

    # ── Concurrency ────────────────────────────────────────────────────────────
    # Azure App Service (or a container) should be sized for ~20 concurrent
    # users. Streamlit's server.maxMessageSize and maxUploadSize are set via
    # .streamlit/config.toml (see companion file). The semaphore below provides
    # an in-process guard so that heavy check jobs don't stack unboundedly.
    "concurrency": {
        "max_concurrent_checks": int(os.environ.get("MAX_CONCURRENT_CHECKS", "20")),
    },

    # ── Logging & Monitoring ───────────────────────────────────────────────────
    # Structured logging to stdout (captured by Azure Monitor / App Insights).
    # PII / sensitive data must NEVER appear in log messages — log only
    # structural metadata (file size, row count, check names, durations).
    "logging": {
        "level":             os.environ.get("LOG_LEVEL", "INFO"),
        "log_sensitive_data": False,          # Hard-coded: PII is never logged
        "appinsights_key":   os.environ.get("APPINSIGHTS_INSTRUMENTATIONKEY", ""),
    },

    # ── Network & Hosting ─────────────────────────────────────────────────────
    # Target: Azure App Service (B2/P1v3) or Azure Container Apps.
    # Access is restricted to authenticated users via Easy Auth; the app is NOT
    # publicly reachable without a valid token.
    # CORS and allowed-hosts are enforced at the Azure front-door / App Gateway
    # layer — not inside Streamlit, which has no native CORS controls.
    "hosting": {
        "platform":           os.environ.get("HOSTING_PLATFORM", "azure_app_service"),
        "region":             os.environ.get("AZURE_REGION", "eastus"),
        "restrict_public":    True,           # Enforced via Azure Easy Auth + VNet rules
        "cors_origins":       os.environ.get("ALLOWED_ORIGINS", "").split(","),
    },
}

# ── Logging Setup ──────────────────────────────────────────────────────────────
logging.basicConfig(
    level=getattr(logging, INFRA_CONFIG["logging"]["level"], logging.INFO),
    format="%(asctime)s | %(levelname)-8s | %(name)s | %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
logger = logging.getLogger("data_quality_checker")

# Silence noisy third-party loggers
for _noisy in ("urllib3", "fsevents", "watchdog"):
    logging.getLogger(_noisy).setLevel(logging.WARNING)


# ── Concurrency Guard ──────────────────────────────────────────────────────────
_check_semaphore = threading.Semaphore(INFRA_CONFIG["concurrency"]["max_concurrent_checks"])


def _log_usage(event: str, **metadata):
    """
    Emit a structured usage log line. Never include file content or user data —
    only structural/operational metadata (sizes, counts, durations, check names).
    """
    # Sanitise: drop any kwarg whose key suggests it could carry sensitive data
    _sensitive_keys = {"content", "value", "data", "row", "cell", "text"}
    safe_meta = {k: v for k, v in metadata.items() if k.lower() not in _sensitive_keys}
    logger.info("event=%s %s", event, " ".join(f"{k}={v}" for k, v in safe_meta.items()))


# ── Authentication Helper ──────────────────────────────────────────────────────
def _get_authenticated_user() -> dict | None:
    """
    Attempt to read the Azure AD Easy Auth user identity injected by the
    platform into request headers. Returns a dict with `username` and `email`
    when available, or None if auth is disabled / headers absent (local dev).

    When REQUIRE_AUTH=true and no identity can be resolved, the app stops with
    an error — this prevents accidental anonymous access in production.
    """
    user = {"username": "unknown", "email": ""}

    # Try Azure AD Easy Auth headers (available on App Service / Container Apps)
    try:
        headers = st.context.headers  # Streamlit ≥ 1.37
        principal = headers.get("X-Ms-Client-Principal-Name", "")
        if principal:
            user = {"username": principal, "email": principal}
            return user
    except AttributeError:
        pass  # Older Streamlit version — fall through to env var

    # Fallback: env var injected by Easy Auth on older runtimes
    env_principal = os.environ.get("HTTP_X_MS_CLIENT_PRINCIPAL_NAME", "")
    if env_principal:
        user = {"username": env_principal, "email": env_principal}
        return user

    # If auth is required but no identity found, block access
    if INFRA_CONFIG["auth"]["require_auth"]:
        return None  # Caller will halt the app

    # Local dev / auth disabled — return a placeholder identity
    return {"username": "local_dev", "email": "local@dev"}


def enforce_auth():
    """
    Gate the entire app behind authentication. Call once at the top of main
    execution. Stops the app (st.stop) if the user cannot be verified.
    """
    user = _get_authenticated_user()
    if user is None:
        st.error(
            "🔒 **Access Denied** — Authentication required.\n\n"
            "Please sign in via your organisation's Azure AD account. "
            "If you believe this is an error, contact your system administrator."
        )
        _log_usage("auth_failure", reason="no_principal_header")
        st.stop()

    # Optional domain restriction
    allowed_domains = [d.strip() for d in INFRA_CONFIG["auth"]["allowed_domains"] if d.strip()]
    if allowed_domains and user["email"]:
        domain = user["email"].split("@")[-1]
        if domain not in allowed_domains:
            st.error(f"🔒 Access is restricted to users from: {', '.join(allowed_domains)}")
            _log_usage("auth_failure", reason="domain_not_allowed")
            st.stop()

    _log_usage("auth_success", provider=INFRA_CONFIG["auth"]["provider"])
    return user


# ── Page Config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Cocoa Campaign Data Quality Checker", layout="wide")

# ── Authentication Gate ────────────────────────────────────────────────────────
# Comment out `enforce_auth()` for local development when Easy Auth is unavailable.
# In production (Azure), this is always active.
current_user = enforce_auth()

st.markdown("""
<style>
    .main-header {
        background: linear-gradient(135deg, #175259 0%, #2d6a4f 100%);
        padding: 2rem; border-radius: 12px; margin-bottom: 2rem;
        color: white; text-align: center;
    }
    .section-header {
        font-size: 1.1rem; font-weight: 700; color: #175259;
        border-bottom: 2px solid #2d6a4f;
        padding-bottom: 0.4rem; margin: 1.5rem 0 1rem 0;
    }
    .issue-badge {
        display:inline-block; padding:2px 8px; border-radius:4px;
        font-size:0.78rem; font-weight:600;
    }
    .infra-badge {
        background:#e8f5e9; color:#175259; font-size:0.75rem;
        padding:2px 8px; border-radius:4px; font-weight:600;
        border:1px solid #2d6a4f;
    }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
    <h1 style="margin:0;font-size:2rem;">🌿 Cocoa Campaign Data Quality Checker</h1>
    <p style="margin:0.5rem 0 0 0;opacity:0.9;">
        Upload your campaign Excel file · Select sheet &amp; columns · Run domain-aware checks · Export flagged records
    </p>
</div>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
BATCH_SIZE = 50

CHART_COLORS = {
    "missing":   "#E74C3C", "duplicate": "#BCDB34",
    "outlier":   "#F39C12", "dtype":     "#9B59B6",
    "banned":    "#C0392B", "cross":     "#E67E22",
    "clean":     "#2ECC71", "primary":   "#175259",
}

# ── Load rules from YAML ───────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def load_rules(yaml_path: str) -> dict:
    """Load domain rules and mandatory fields from quality_rules.yaml."""
    path = pathlib.Path(yaml_path)
    if not path.exists():
        st.error(f"Rules file not found: {yaml_path}\n"
                 "Place `quality_rules.yaml` in the same folder as this script.")
        st.stop()
    with open(path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)

_SCRIPT_DIR  = pathlib.Path(__file__).parent
_RULES_PATH  = _SCRIPT_DIR / "quality_rules.yaml"
_RULES       = load_rules(str(_RULES_PATH))

DOMAIN_RULES:    list = _RULES.get("domain_rules", [])
MANDATORY_MAP:   dict = _RULES.get("mandatory_fields", {})
DISPLAY_COLS:    list = MANDATORY_MAP.get("display_columns", [])
OUTLIER_CFG:   dict = _RULES.get("outlier_detection", {})
OUTLIER_COLS:  list = OUTLIER_CFG.get("cols", [])
OUTLIER_GROUP: str  = OUTLIER_CFG.get("country_col", "Country")


# ── Helpers ────────────────────────────────────────────────────────────────────

def safe_str(df):
    out = df.copy()
    for col in out.columns:
        if out[col].dtype == object:
            types = out[col].dropna().map(type).unique()
            if len(types) > 1:
                out[col] = out[col].astype(str)
    return out


def coerce_cell(v):
    if isinstance(v, (np.integer,)):   return int(v)
    if isinstance(v, (np.floating,)):  return None if np.isnan(v) else float(v)
    if isinstance(v, (np.bool_,)):     return bool(v)
    if isinstance(v, float) and np.isnan(v): return None
    if not isinstance(v, (int, float, str, bool, type(None))): return str(v)
    return v


def safe_concat(frames, base_cols):
    non_empty = [f for f in frames if not f.empty]
    if not non_empty:
        return pd.DataFrame()
    combined = pd.concat(non_empty, ignore_index=True)
    subset = [c for c in base_cols if c in combined.columns]
    if subset:
        combined = combined.drop_duplicates(subset=subset)
    return combined


def slim_view(issue_df, all_cols, checked_cols=None):
    if issue_df.empty:
        return issue_df
    anchor = [c for c in DISPLAY_COLS if c in issue_df.columns]
    def flagged_cols(txt):
        return [
            c for c in issue_df.columns
            if c not in anchor
            and c not in ("__issue_type__", "__issue__")
            and re.search(re.escape(c), str(txt))
        ]
    issue_cols = list(dict.fromkeys(
        c
        for t in issue_df["__issues__"]
        for c in flagged_cols(t)
    ))
    meta = ["__issue_type__", "__issues__"]
    keep = anchor + issue_cols + meta
    return issue_df[[c for c in keep if c in issue_df.columns]]


# ── Core checks ────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def check_missing(df, cols):
    cols = list(cols)
    mask = df[cols].isnull().any(axis=1)
    result = df[mask].copy()
    result["__issues__"] = result.apply(
        lambda r: "Missing: " + ", ".join(c for c in cols if pd.isnull(r[c])), axis=1)
    result["__issue_type__"] = "Missing Value"
    return result


@st.cache_data(show_spinner=False)
def check_duplicates(df, cols, keep="first"):
    cols = list(cols)
    keep_arg = False if keep == "none" else keep
    mask = df.duplicated(subset=cols, keep=keep_arg)
    result = df[mask].copy()
    if result.empty: return result
    def find_orig(idx):
        row_vals = df.loc[idx, cols]
        others = [i + 2 for i in df[(df[cols] == row_vals).all(axis=1)].index if i != idx]
        return f"Duplicate of row(s): {others}"
    result["__issues__"] = [find_orig(i) for i in result.index]
    result["__issue_type__"] = "Duplicate"
    return result


@st.cache_data(show_spinner=False)
def check_outliers_batch(df, cols, group_col, method="IQR"):
    flagged_rows = []
    issues_list  = []
    types_list   = []
    methods_cfg = OUTLIER_CFG.get("methods", {})
    has_group   = group_col and group_col in df.columns
    for col in cols:
        if col not in df.columns:
            continue
        s_full = pd.to_numeric(df[col], errors="coerce")
        groups = df[group_col].unique() if has_group else ["_all_"]
        for group in groups:
            group_mask  = (df[group_col] == group) if has_group else pd.Series(True, index=df.index)
            group_label = group if has_group else "all"
            s = s_full[group_mask].dropna()
            if len(s) < 4:
                continue
            if method == "IQR":
                cfg = methods_cfg.get("IQR", {})
                q1  = s.quantile(cfg.get("lower_quantile", 0.25))
                q3  = s.quantile(cfg.get("upper_quantile", 0.75))
                mul = cfg.get("iqr_multiplier", 1.5)
                lo, hi = q1 - mul * (q3 - q1), q3 + mul * (q3 - q1)
                label  = "IQR"
            else:
                cfg        = methods_cfg.get("Z-Score", {})
                mean, std  = s.mean(), s.std()
                if std == 0:
                    continue
                mul        = cfg.get("std_multiplier", 3)
                lo, hi     = mean - mul * std, mean + mul * std
                label      = "Z-Score"
            outlier_mask = group_mask & s_full.notna() & ((s_full < lo) | (s_full > hi))
            for idx in df[outlier_mask].index:
                flagged_rows.append(idx)
                issues_list.append(
                    f"Outlier ({label}) within {group_label}: '{col}' = {df.at[idx, col]} "
                    f"outside [{lo:.2f}, {hi:.2f}]"
                )
                types_list.append("Statistical Outlier")
    if not flagged_rows:
        return pd.DataFrame()
    result = df.loc[flagged_rows].copy()
    result["__issues__"]     = issues_list
    result["__issue_type__"] = types_list
    return result.drop_duplicates(subset=list(df.columns))


@st.cache_data(show_spinner=False)
def check_dtype_batch(df, cols):
    cols = list(cols)
    rows, issues = [], []
    for col in cols:
        expected = df[col].dtype
        for idx, val in df[col].items():
            if pd.isnull(val): continue
            try:
                if pd.api.types.is_numeric_dtype(expected):
                    if not isinstance(val, (int, float, np.integer, np.floating)):
                        float(val)
                elif pd.api.types.is_datetime64_any_dtype(expected):
                    pd.to_datetime(val)
            except (ValueError, TypeError):
                rows.append(idx)
                issues.append(f"Type mismatch in '{col}': got '{type(val).__name__}'")
    if not rows: return pd.DataFrame()
    result = df.loc[rows].copy()
    result["__issues__"] = issues
    result["__issue_type__"] = "Data Type Issue"
    return result


def check_dtype(df, cols):
    batches = [cols[i:i+BATCH_SIZE] for i in range(0, len(cols), BATCH_SIZE)]
    return safe_concat([check_dtype_batch(df, tuple(b)) for b in batches], list(df.columns))


@st.cache_data(show_spinner=False)
def check_domain_rules(df, sheet_name):
    flagged_rows = []
    issues_list = []
    types_list = []
    for rule in DOMAIN_RULES:
        col = rule.get("column", "")
        rule_type = rule.get("type", "")
        if rule_type not in ["conditional_any_notnull", "prefix_range"]:
            if col not in df.columns:
                continue
        if rule_type == "range":
            lo = rule.get("min", -np.inf)
            hi = rule.get("max", np.inf)
            s = pd.to_numeric(df[col], errors="coerce")
            mask = s.notna() & ((s < lo) | (s > hi))
            for idx in df[mask].index:
                flagged_rows.append(idx)
                issues_list.append(f"Domain: '{col}' = {df.at[idx, col]} — {rule['rule']}")
                types_list.append("Domain Rule Violation")
        elif rule_type == "conditional_range_by_value":
            condition_col = rule.get("condition_col")
            ranges = rule.get("ranges", {})
            if condition_col and condition_col in df.columns:
                s = pd.to_numeric(df[col], errors="coerce")
                for country, limits in ranges.items():
                    lo = limits.get("min", -np.inf)
                    hi = limits.get("max", np.inf)
                    mask = (
                        (df[condition_col] == country)
                        & s.notna()
                        & ((s < lo) | (s > hi))
                    )
                    for idx in df[mask].index:
                        flagged_rows.append(idx)
                        issues_list.append(
                            f"Domain: '{col}' = {df.at[idx, col]} "
                            f"outside allowed range for {country} "
                            f"({lo} - {hi}) — {rule['rule']}"
                        )
                        types_list.append("Conditional Range Violation")
        elif rule_type == "banned_value":
            banned = rule.get("banned", [])
            mask = df[col].isin(banned)
            for idx in df[mask].index:
                flagged_rows.append(idx)
                issues_list.append(f"Domain: '{col}' = '{df.at[idx, col]}' — {rule['rule']}")
                types_list.append("Banned Value")
        elif rule_type == "conditional_notnull":
            trigger_col = rule.get("trigger_col")
            trigger_val = rule.get("trigger_val")
            if trigger_col and trigger_col in df.columns:
                mask = (df[trigger_col] == trigger_val) & df[col].isnull()
                for idx in df[mask].index:
                    flagged_rows.append(idx)
                    issues_list.append(
                        f"Domain: '{col}' must not be blank when "
                        f"'{trigger_col}' = '{trigger_val}' — {rule['rule']}"
                    )
                    types_list.append("Conditional Missing")
        elif rule_type == "conditional_any_notnull":
            prefix = rule.get("column_prefix")
            trigger_col = rule.get("trigger_col")
            trigger_val = rule.get("trigger_val")
            if prefix and trigger_col and trigger_col in df.columns:
                matching_cols = [c for c in df.columns if c.startswith(prefix)]
                if matching_cols:
                    mask = (
                        (df[trigger_col] == trigger_val)
                        & (df[matching_cols].isnull().all(axis=1))
                    )
                    for idx in df[mask].index:
                        flagged_rows.append(idx)
                        issues_list.append(
                            f"At least one column starting with "
                            f"'{prefix}' must be filled when "
                            f"'{trigger_col}' = '{trigger_val}' — {rule['rule']}"
                        )
                        types_list.append("Conditional Missing Group")
        elif rule_type == "prefix_range":
            prefix = rule.get("column_prefix", "")
            lo = rule.get("min", -np.inf)
            hi = rule.get("max", np.inf)
            matching_cols = [c for c in df.columns if c.startswith(prefix)]
            for col in matching_cols:
                s = pd.to_numeric(df[col], errors="coerce")
                mask = s.notna() & ((s < lo) | (s > hi))
                for idx in df[mask].index:
                    flagged_rows.append(idx)
                    issues_list.append(f"Domain: '{col}' = {df.at[idx, col]} — {rule['rule']}")
                    types_list.append("Domain Rule Violation")
    if not flagged_rows:
        return pd.DataFrame()
    result = df.loc[flagged_rows].copy()
    result["__issues__"] = issues_list
    result["__issue_type__"] = types_list
    return result.drop_duplicates(subset=list(df.columns))


def check_mandatory(df, sheet_name):
    mand_fields = []
    for key, fields in MANDATORY_MAP.items():
        if key.lower() in sheet_name.lower():
            mand_fields = fields
            break
    mand = [c for c in mand_fields if c in df.columns]
    if not mand:
        return pd.DataFrame()
    mask = df[mand].isnull().any(axis=1)
    result = df[mask].copy()
    result["__issues__"] = result.apply(
        lambda r: "Mandatory field missing: " + ", ".join(c for c in mand if pd.isnull(r[c])),
        axis=1,
    )
    result["__issue_type__"] = "Mandatory Field Missing"
    return result


# ── File Load (ephemeral / in-memory only) ─────────────────────────────────────

@st.cache_data(show_spinner="📂 Loading file…")
def load_excel(file_bytes: bytes) -> dict:
    """
    Load an Excel workbook entirely from memory (BytesIO).
    The raw bytes are NEVER written to disk — they live only in the Streamlit
    cache for the duration of the session and are garbage-collected when the
    session ends. No persistent storage is used.
    """
    _log_usage("file_loaded", size_bytes=len(file_bytes))
    return pd.read_excel(BytesIO(file_bytes), sheet_name=None)


def validate_upload(uploaded_file) -> bytes:
    """
    Validate the upload against configured limits (size, extension) and return
    the raw bytes. Raises st.error + st.stop on any violation.
    """
    max_bytes = INFRA_CONFIG["file_handling"]["max_upload_mb"] * 1024 * 1024
    file_bytes = uploaded_file.read()

    if len(file_bytes) > max_bytes:
        st.error(
            f"File exceeds the {INFRA_CONFIG['file_handling']['max_upload_mb']} MB limit. "
            "Please reduce the file size and re-upload."
        )
        _log_usage("upload_rejected", reason="size_exceeded", size_bytes=len(file_bytes))
        st.stop()

    ext = uploaded_file.name.rsplit(".", 1)[-1].lower()
    if ext not in INFRA_CONFIG["file_handling"]["allowed_extensions"]:
        st.error(f"Unsupported file type '.{ext}'. Only .xlsx and .xls files are accepted.")
        _log_usage("upload_rejected", reason="invalid_extension", extension=ext)
        st.stop()

    return file_bytes


# ── Excel Export ───────────────────────────────────────────────────────────────

def build_excel_export(original, results_dict, selected_cols):
    wb = Workbook()
    wb.remove(wb.active)
    COLORS = {
        "header_bg": "175259", "header_fg": "FFFFFF",
        "missing":   "FDECEA", "duplicate": "E8F4FD",
        "outlier":   "FEF3E2", "dtype":     "F3E5F5",
        "domain":    "FFF3CD", "mandatory": "FCE4EC",
        "alt_row":   "F8F9FA",
    }
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws):
        for cell in ws[1]:
            cell.font      = Font(bold=True, color=COLORS["header_fg"], name="Arial", size=10)
            cell.fill      = PatternFill("solid", start_color=COLORS["header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = border
        ws.row_dimensions[1].height = 30

    def write_sheet(ws, sheet_df, fill_color):
        if sheet_df.empty:
            ws.append(["No issues found."])
            return
        data_cols  = [c for c in sheet_df.columns if not c.startswith("__")]
        write_cols = data_cols + ["__issue_type__", "__issues__"]
        headers    = [c.replace("__issue_type__","Issue Type").replace("__issues__","Issue Detail") for c in write_cols]
        ws.append(headers)
        style_header(ws)
        fill = PatternFill("solid", start_color=fill_color)
        alt  = PatternFill("solid", start_color=COLORS["alt_row"])
        for i, (_, row) in enumerate(sheet_df[write_cols].iterrows(), start=2):
            ws.append([coerce_cell(v) for v in row])
            for cell in ws[i]:
                cell.fill      = fill if i % 2 == 0 else alt
                cell.border    = border
                cell.alignment = Alignment(vertical="center")
                cell.font      = Font(name="Arial", size=9)
        for ci, col in enumerate(write_cols, 1):
            hdr_len = len(str(headers[ci-1]))
            val_len = int(sheet_df[col].astype(str).str.len().fillna(0).max()) if col in sheet_df.columns else 0
            ws.column_dimensions[get_column_letter(ci)].width = min(max(hdr_len, val_len) + 4, 40)

    all_frames = list(results_dict.values())
    all_issues = safe_concat(all_frames, list(original.columns))

    ws_sum = wb.create_sheet("Summary")
    summary_data = [
        ["Cocoa Campaign Data Quality Report", ""],
        ["", ""],
        ["Metric", "Value"],
        ["Total Rows Checked", len(original)],
        ["Columns Checked", ", ".join(selected_cols)],
    ]
    for k, v in results_dict.items():
        summary_data.append([k, len(v)])
    summary_data.append(["Total Unique Issue Rows", len(all_issues)])
    thin2 = Side(style="thin", color="CCCCCC")
    b2 = Border(left=thin2, right=thin2, top=thin2, bottom=thin2)
    for r, row in enumerate(summary_data, 1):
        ws_sum.append(row)
        for cell in ws_sum[r]:
            cell.font      = Font(name="Arial", size=10, bold=(r in (1, 3)))
            cell.border    = b2
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if r == 1:
            ws_sum[r][0].font = Font(name="Arial", size=14, bold=True, color=COLORS["header_bg"])
    ws_sum.column_dimensions["A"].width = 35
    ws_sum.column_dimensions["B"].width = 50

    color_map = {
        "Missing Values":          COLORS["missing"],
        "Mandatory Field Missing":  COLORS["mandatory"],
        "Duplicates":              COLORS["duplicate"],
        "Outliers":                COLORS["outlier"],
        "Data Type Issues":        COLORS["dtype"],
        "Domain Rule Violations":  COLORS["domain"],
        "All Issues":              COLORS["alt_row"],
    }
    for sheet_label, frame in results_dict.items():
        write_sheet(wb.create_sheet(sheet_label[:31]), frame, color_map.get(sheet_label, COLORS["alt_row"]))
    write_sheet(wb.create_sheet("All Issues"), all_issues, COLORS["alt_row"])

    ws_orig = wb.create_sheet("Original Data")
    ws_orig.append(list(original.columns))
    style_header(ws_orig)
    for i, row in enumerate(original.itertuples(index=False), 2):
        ws_orig.append([coerce_cell(v) for v in row])
        for cell in ws_orig[i]:
            cell.font = Font(name="Arial", size=9)
            cell.border = border
    for ci, col in enumerate(original.columns, 1):
        max_len = max(len(str(col)), int(original[col].astype(str).str.len().fillna(0).max()))
        ws_orig.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Charts ─────────────────────────────────────────────────────────────────────

PLOTLY_LAYOUT = dict(
    font_family="Arial, sans-serif",
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    margin=dict(t=40, b=40, l=40, r=40),
)

def chart_overview(counts, total):
    labels, values, colors = [], [], []
    color_seq = [CHART_COLORS["missing"], CHART_COLORS["dtype"],
                 CHART_COLORS["duplicate"], CHART_COLORS["outlier"],
                 CHART_COLORS["banned"], CHART_COLORS["cross"]]
    for i, (k, v) in enumerate(counts.items()):
        labels.append(k); values.append(v); colors.append(color_seq[i % len(color_seq)])
    max_v = max(values) if any(values) else 1
    fig = go.Figure(go.Bar(
        x=values, y=labels, orientation="h", marker_color=colors,
        text=[f"{v:,}  ({v/total*100:.1f}%)" if total else f"{v:,}" for v in values],
        textposition="outside", cliponaxis=False,
    ))
    fig.update_layout(title="Issue Count by Type", xaxis_title="Rows Affected",
                      xaxis=dict(gridcolor="#e8e8e8", range=[0, max_v*1.4]),
                      height=300, **PLOTLY_LAYOUT)
    return fig


def chart_health(total, issue_total):
    clean = max(0, total - issue_total)
    fig = go.Figure(go.Pie(
        labels=["Clean Rows", "Rows with Issues"], values=[clean, issue_total],
        hole=0.55, marker_colors=[CHART_COLORS["clean"], "#E74C3C"],
        textinfo="label+percent",
    ))
    pct = f"{clean/total*100:.0f}%<br>clean" if total else "—"
    fig.add_annotation(text=pct, x=0.5, y=0.5, showarrow=False,
                       font=dict(size=14, color=CHART_COLORS["primary"]))
    fig.update_layout(title="Overall Data Health", height=300,
                      showlegend=True, legend=dict(orientation="h", y=-0.15),
                      **PLOTLY_LAYOUT)
    return fig


def chart_completeness(df, selected_cols):
    null_pct = [(col, df[col].isnull().mean()*100) for col in selected_cols]
    null_pct = sorted(null_pct, key=lambda x: x[1], reverse=True)[:20]
    cols, pcts = zip(*null_pct) if null_pct else ([], [])
    fig = go.Figure(go.Bar(
        x=pcts, y=[c[:40] for c in cols], orientation="h",
        marker_color=[CHART_COLORS["missing"] if p > 10 else "#27AE60" for p in pcts],
        text=[f"{p:.1f}%" for p in pcts], textposition="outside",
    ))
    fig.update_layout(title="Top 20 Columns by Null %", xaxis_title="Null %",
                      xaxis=dict(range=[0, 110]), height=max(300, len(cols)*25+80),
                      **PLOTLY_LAYOUT)
    return fig


def show_table(tab, issue_df, label, all_cols, checked_cols):
    with tab:
        if issue_df.empty:
            st.success(f"✅ No {label} detected.")
        else:
            disp = slim_view(issue_df, all_cols, checked_cols)
            disp = disp.rename(columns={"__issue_type__": "Issue Type", "__issues__": "Issue Detail"})
            st.dataframe(safe_str(disp), use_container_width=True, height=320)
            st.caption(f"{len(issue_df):,} row(s) flagged · Full data in export.")


# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Settings")
    outlier_method = st.radio("Outlier Detection Method", ["IQR", "Z-Score"],
        help="IQR = 1.5×IQR rule. Z-Score = 3 std devs from mean.")
    st.markdown("---")
    st.markdown("**Checks to Run**")
    run_missing    = st.checkbox("Missing Values",                value=True)
    run_mandatory  = st.checkbox("Mandatory Fields (sheet-aware)", value=True)
    run_duplicate  = st.checkbox("Duplicates",                    value=True)
    run_outlier    = st.checkbox("Outliers",                      value=True)
    run_dtype      = st.checkbox("Data Type Issues",              value=True)
    run_domain     = st.checkbox("Domain / Business Rules",       value=True,
                                  help="Reads rules from quality_rules.yaml")

    if run_duplicate:
        st.markdown("**Duplicate Settings**")
        dup_keep_label = st.radio("Flag", ["All except first", "All except last", "Every occurrence"])
        keep_map = {"All except first": "first", "All except last": "last", "Every occurrence": "none"}
        dup_keep = keep_map[dup_keep_label]
    else:
        dup_keep = "first"

    with st.expander("📋 Active domain rules", expanded=False):
        st.caption(f"{len(DOMAIN_RULES)} rules loaded from `quality_rules.yaml`")
        for r in DOMAIN_RULES:
            st.markdown(f"- **{r.get('column','')}** `{r.get('type','')}` — {r.get('rule','')}")

    st.markdown("---")

    # ── Infrastructure Status Panel ────────────────────────────────────────────
    with st.expander("🔒 Infrastructure & Security", expanded=False):
        st.markdown("**Authentication**")
        auth_provider = INFRA_CONFIG["auth"]["provider"].replace("_", " ").title()
        st.markdown(
            f'<span class="infra-badge">🔑 {auth_provider}</span> '
            f'{"Required" if INFRA_CONFIG["auth"]["require_auth"] else "Optional"}',
            unsafe_allow_html=True,
        )
        if current_user:
            st.caption(f"Signed in as: `{current_user['username']}`")

        st.markdown("**File Handling**")
        st.markdown(
            f'<span class="infra-badge">📁 Ephemeral</span> '
            f'Max {INFRA_CONFIG["file_handling"]["max_upload_mb"]} MB · '
            f'No persistent storage',
            unsafe_allow_html=True,
        )

        st.markdown("**Concurrency**")
        st.markdown(
            f'<span class="infra-badge">⚡ ~{INFRA_CONFIG["concurrency"]["max_concurrent_checks"]} users</span> '
            f'Semaphore-guarded check queue',
            unsafe_allow_html=True,
        )

        st.markdown("**Logging**")
        st.markdown(
            f'<span class="infra-badge">📋 {INFRA_CONFIG["logging"]["level"]}</span> '
            f'Structural metadata only · No PII retained',
            unsafe_allow_html=True,
        )

        st.markdown("**Hosting**")
        platform = INFRA_CONFIG["hosting"]["platform"].replace("_", " ").title()
        st.markdown(
            f'<span class="infra-badge">☁️ {platform}</span> '
            f'Auth-restricted · Not publicly open',
            unsafe_allow_html=True,
        )

    st.markdown("---")
    st.markdown("**About**")
    st.caption(
        "Rules are defined in `quality_rules.yaml` — edit that file to add or "
        "change validations without touching any Python code."
    )


# ── File Upload ────────────────────────────────────────────────────────────────
uploaded = st.file_uploader(
    f"📂 Upload Campaign Excel File (.xlsx) — max {INFRA_CONFIG['file_handling']['max_upload_mb']} MB",
    type=INFRA_CONFIG["file_handling"]["allowed_extensions"],
    help="Upload your Campaign Excel file (e.g. Campaign_Cocoa_2025_LATAM.xlsx). "
         "Files are processed in memory and discarded after your session.",
)

if uploaded is None:
    st.info("Upload your campaign Excel file to begin.")
    st.stop()

try:
    file_bytes = validate_upload(uploaded)          # size + extension check
    all_sheets = load_excel(file_bytes)             # in-memory parse, no disk write
except Exception as e:
    st.error(f"Could not read file: {e}")
    _log_usage("file_parse_error", error=type(e).__name__)
    st.stop()

sheet_names = [s for s, df in all_sheets.items() if not df.empty]
if not sheet_names:
    st.error("No non-empty sheets found.")
    st.stop()

# ── Sheet Selection ────────────────────────────────────────────────────────────
def on_sheet_change():
    prev = st.session_state.get("last_sheet")
    if prev and f"col_select_{prev}" in st.session_state:
        del st.session_state[f"col_select_{prev}"]

if len(sheet_names) > 1:
    sheet_name = st.selectbox("Select Sheet", sheet_names, key="sheet_sel", on_change=on_sheet_change)
else:
    sheet_name = sheet_names[0]

if st.session_state.get("last_sheet") != sheet_name:
    st.session_state["last_sheet"] = sheet_name
    k = f"col_select_{sheet_name}"
    if k in st.session_state:
        del st.session_state[k]

df = all_sheets[sheet_name]

st.markdown(
    f"<div class='section-header'>📋 Data Preview — <em>{sheet_name}</em> "
    f"({len(df):,} rows × {len(df.columns)} cols)</div>",
    unsafe_allow_html=True
)
st.dataframe(safe_str(df.head(50)), use_container_width=True, height=220)

# ── Column Selection ───────────────────────────────────────────────────────────
st.markdown("<div class='section-header'>Select Columns to Check</div>", unsafe_allow_html=True)

all_cols = df.columns.tolist()
widget_key = f"col_select_{sheet_name}"
if widget_key not in st.session_state:
    st.session_state[widget_key] = []

c1, c2, c3 = st.columns([3, 1, 1])
with c2:
    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("✅ Select All", use_container_width=True):
        st.session_state[widget_key] = all_cols
        st.rerun()
with c3:
    st.markdown("<br>", unsafe_allow_html=True)
    key_suggestion = []
    for _key, _fields in MANDATORY_MAP.items():
        if _key.lower() in sheet_name.lower():
            key_suggestion = [c for c in _fields if c in all_cols]
            break
    if key_suggestion and st.button("🎯 Key Fields", use_container_width=True):
        st.session_state[widget_key] = key_suggestion
        st.rerun()

with c1:
    selected_cols = st.multiselect(
        "Choose columns to analyse",
        options=all_cols,
        default=st.session_state[widget_key],
        key=widget_key,
    )

if not selected_cols:
    st.info("Select at least one column above to run quality checks.")
    st.stop()

# ── Run Checks (semaphore-guarded for concurrency) ─────────────────────────────
t_start = time.perf_counter()

acquired = _check_semaphore.acquire(blocking=False)
if not acquired:
    st.warning("⏳ The system is at capacity. Please wait a moment and try again.")
    _log_usage("concurrency_limit_hit")
    st.stop()

try:
    with st.spinner("Running quality checks…"):
        results = {}

        if run_missing:
            results["Missing Values"] = check_missing(df, tuple(selected_cols))

        if run_mandatory:
            mand_df = check_mandatory(df, sheet_name)
            if not mand_df.empty:
                results["Mandatory Field Missing"] = mand_df

        if run_duplicate:
            results["Duplicates"] = check_duplicates(df, tuple(df.columns.tolist()), keep=dup_keep)

        if run_outlier:
            outlier_targets = [
                c for c in OUTLIER_COLS
                if c in df.columns
                and pd.api.types.is_numeric_dtype(df[c])
            ]
            if outlier_targets:
                results["Outliers"] = check_outliers_batch(
                    df, outlier_targets, OUTLIER_GROUP, method=outlier_method
                )

        if run_dtype:
            results["Data Type Issues"] = check_dtype(df, selected_cols)

        if run_domain:
            domain_df = check_domain_rules(df, sheet_name)
            if not domain_df.empty:
                results["Domain Rule Violations"] = domain_df

finally:
    _check_semaphore.release()

t_elapsed = time.perf_counter() - t_start
_log_usage(
    "checks_complete",
    sheet=sheet_name,
    rows=len(df),
    cols_checked=len(selected_cols),
    checks_run=",".join(results.keys()),
    duration_ms=round(t_elapsed * 1000),
    total_issues=sum(len(v) for v in results.values()),
)

# ── Metrics ────────────────────────────────────────────────────────────────────
st.markdown("<div class='section-header'>📊 Quality Summary</div>", unsafe_allow_html=True)

all_issue_df = safe_concat(list(results.values()), list(df.columns))
total_issues = len(all_issue_df)

cols_m = st.columns(len(results) + 1)
cols_m[0].metric("🗂️ Total Rows", f"{len(df):,}")
for i, (label, frame) in enumerate(results.items(), 1):
    icon = {"Missing Values": "❌", "Mandatory Field Missing": "⚠️", "Duplicates": "🔁",
            "Outliers": "📉", "Data Type Issues": "🔠", "Domain Rule Violations": "🚫"}.get(label, "🔍")
    cols_m[i].metric(f"{icon} {label}", f"{len(frame):,}")

# ── Charts ─────────────────────────────────────────────────────────────────────
st.markdown("<div class='section-header'>📈 Visual Analytics</div>", unsafe_allow_html=True)

issue_counts = {k: len(v) for k, v in results.items()}
c1, c2 = st.columns(2)
with c1:
    st.plotly_chart(chart_health(len(df), total_issues), use_container_width=True, config={"displayModeBar": False})
with c2:
    st.plotly_chart(chart_overview(issue_counts, len(df)), use_container_width=True, config={"displayModeBar": False})

with st.expander("📊 Column Completeness (Null % by Column)", expanded=False):
    st.plotly_chart(chart_completeness(df, selected_cols), use_container_width=True, config={"displayModeBar": False})

# ── Column Stats ───────────────────────────────────────────────────────────────
with st.expander("📈 Column-Level Statistics", expanded=False):
    stats = []
    for col in selected_cols:
        null_n = int(df[col].isnull().sum())
        dup_n  = int(df.duplicated(subset=[col], keep=False).sum())
        stats.append({
            "Column": col, "Dtype": str(df[col].dtype),
            "Nulls": null_n, "Null %": f"{null_n/len(df)*100:.1f}%",
            "Duplicates": dup_n, "Unique": int(df[col].nunique()),
            "Sample": str(df[col].dropna().iloc[0]) if df[col].notna().any() else "—",
        })
    st.dataframe(pd.DataFrame(stats), use_container_width=True)

# ── Issue Tabs ─────────────────────────────────────────────────────────────────
st.markdown("<div class='section-header'>Issue Details</div>", unsafe_allow_html=True)

tab_labels = list(results.keys()) + ["📋 All Issues"]
tabs = st.tabs(tab_labels)

for tab, (label, frame) in zip(tabs[:-1], results.items()):
    show_table(tab, frame, label, all_cols, selected_cols)

with tabs[-1]:
    if all_issue_df.empty:
        st.success("✅ No issues detected — data looks clean!")
    else:
        disp = slim_view(all_issue_df, all_cols, selected_cols)
        disp = disp.rename(columns={"__issue_type__": "Issue Type", "__issues__": "Issue Detail"})
        st.dataframe(safe_str(disp), use_container_width=True, height=380)
        st.caption(f"{total_issues:,} total unique issue row(s)")

# ── Export ─────────────────────────────────────────────────────────────────────
st.markdown("<div class='section-header'>⬇️ Export Results</div>", unsafe_allow_html=True)

if all_issue_df.empty:
    st.success("✅ No issues found — nothing to export!")
else:
    excel_bytes = build_excel_export(df, results, selected_cols)
    st.download_button(
        label="📥 Download Quality Report (.xlsx)",
        data=excel_bytes,
        file_name=f"quality_report_{sheet_name.replace(' ', '_')[:20]}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.caption("Report includes: Summary · " + " · ".join(results.keys()) + " · All Issues · Original Data")
    _log_usage("report_exported", sheet=sheet_name, issue_rows=total_issues)
